"""J.5 / J.6 Excel extraction (openpyxl + LibreOffice legacy)."""
from __future__ import annotations

import asyncio
import io
import logging
import random
from collections import Counter
from collections.abc import Iterator
from datetime import date, datetime
from decimal import Decimal
from numbers import Real
from typing import Any

import openpyxl

from trimcp.extractors.common import (
    is_zip_encrypted_ooxml,
    rows_to_markdown,
    split_tables_by_blank_rows,
    trim_trailing_empty,
)
from trimcp.extractors.core import ExtractionResult, Section, empty_skipped
from trimcp.extractors.libreoffice import libreoffice_convert

log = logging.getLogger(__name__)

SHEET_ROW_SMALL = 1000
SHEET_ROW_SAMPLE = 10_000


def _row_count(ws: Any) -> int:
    return sum(1 for _ in ws.iter_rows(values_only=True))


def _collect_rows(ws: Any) -> list[tuple[Any, ...]]:
    return list(ws.iter_rows(values_only=True))


def _format_val(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, (datetime, date)):
        return v.isoformat()
    if isinstance(v, Decimal):
        return format(v, "f")
    return str(v)


def _sheet_comment_lines(ws: Any, max_rows_scan: int = 50_000) -> list[str]:
    lines: list[str] = []
    try:
        for i, row in enumerate(ws.iter_rows()):
            if i > max_rows_scan:
                break
            for cell in row:
                if cell.comment:
                    lines.append(
                        f"[cell {cell.coordinate} comment by {cell.comment.author or '?'}: {cell.comment.text}]"
                    )
    except Exception as e:
        log.debug("cell_comments_skip: %s", e)
    return lines


def _numeric_type(v: Any) -> bool:
    return isinstance(v, Real) and not isinstance(v, bool)


def _summarize_large_sheet(ws: Any, sheet_name: str) -> tuple[str, list[str]]:
    warnings: list[str] = []
    it: Iterator[tuple[Any, ...]] = ws.iter_rows(values_only=True)
    header = next(it, None)
    if not header:
        return "(empty sheet)", warnings
    ncol = len(header)
    row_count = 0
    types_n: list[Counter[str]] = [Counter() for _ in range(ncol)]
    numeric_min: list[float | None] = [None] * ncol
    numeric_max: list[float | None] = [None] * ncol
    str_top: list[Counter[str]] = [Counter() for _ in range(ncol)]
    for row in it:
        row_count += 1
        for j in range(ncol):
            v = row[j] if j < len(row) else None
            if v is None or (isinstance(v, str) and not v.strip()):
                types_n[j]["empty"] += 1
                continue
            if _numeric_type(v):
                types_n[j]["numeric"] += 1
                fv = float(v)
                numeric_min[j] = fv if numeric_min[j] is None else min(numeric_min[j], fv)
                numeric_max[j] = fv if numeric_max[j] is None else max(numeric_max[j], fv)
            else:
                types_n[j]["text"] += 1
                s = _format_val(v)[:200]
                str_top[j][s] += 1
    lines = [
        f"# Sheet: {sheet_name} (summary — full content not indexed; {row_count + 1} rows)",
        f"Columns ({ncol}): " + " | ".join(_format_val(h) for h in header),
    ]
    for j in range(ncol):
        cn = header[j] if j < len(header) else f"col{j}"
        parts = [f"**{_format_val(cn)}**"]
        parts.append(f"types: {dict(types_n[j])}")
        if numeric_min[j] is not None:
            parts.append(f"min={numeric_min[j]}, max={numeric_max[j]}")
        top = str_top[j].most_common(5)
        if top:
            parts.append("top_values: " + "; ".join(f"{k!r} ({c})" for k, c in top))
        lines.append("- " + "; ".join(parts))
    warnings.append(
        f"sheet_summary_only:{sheet_name}: data sheet — full content not indexed for semantic search; query columns by name"
    )
    return "\n".join(lines), warnings


def _emit_sheet_sections(
    sheet_name: str,
    rows: list[tuple[Any, ...]],
    order: int,
    warnings: list[str],
) -> tuple[list[Section], int]:
    sections: list[Section] = []
    nrows = len(rows)
    if nrows > SHEET_ROW_SAMPLE:
        return sections, order
    subtables = split_tables_by_blank_rows(rows) if nrows > 50 else [rows]
    for ti, tbl in enumerate(subtables):
        tbl = trim_trailing_empty(tbl)
        if not tbl:
            continue
        md = rows_to_markdown(tbl)
        path = f"Sheet: {sheet_name}"
        if len(subtables) > 1:
            path = f"{path} > Table {ti + 1}"
        sections.append(Section(text=md, structure_path=path, section_type="sheet", order=order))
        order += 1
    return sections, order


def extract_xlsx_sync(blob: bytes) -> ExtractionResult:
    if is_zip_encrypted_ooxml(blob):
        return empty_skipped("openpyxl", "encrypted")
    warnings: list[str] = []
    try:
        wb = openpyxl.load_workbook(io.BytesIO(blob), read_only=True, data_only=True)
    except Exception as e:
        log.warning("openpyxl_failed: %s", e)
        return empty_skipped("openpyxl", "corrupt", warnings=[str(e)])

    metadata: dict[str, Any] = {"sheet_count": len(wb.sheetnames)}
    try:
        props = wb.properties
        metadata.update(
            {
                "creator": getattr(props, "creator", None),
                "created": props.created.isoformat() if getattr(props, "created", None) else None,
                "modified": props.modified.isoformat() if getattr(props, "modified", None) else None,
            }
        )
    except Exception:
        pass

    sections: list[Section] = []
    order = 0
    for sheet_name in wb.sheetnames:
        try:
            ws = wb[sheet_name]
        except Exception as e:
            warnings.append(f"sheet_open_failed:{sheet_name}:{e}")
            continue
        if getattr(ws, "sheet_state", None) == "hidden":
            warnings.append(f"Skipped hidden sheet: {sheet_name}")
            continue
        nrows = _row_count(ws)
        if nrows == 0:
            continue

        comment_extra: list[str] = []
        if nrows <= SHEET_ROW_SAMPLE:
            comment_extra = _sheet_comment_lines(ws)

        if nrows <= SHEET_ROW_SMALL:
            rows = _collect_rows(ws)
            rows = trim_trailing_empty(rows)
            secs, order = _emit_sheet_sections(sheet_name, rows, order, warnings)
            sections.extend(secs)
            for line in comment_extra:
                sections.append(
                    Section(text=line, structure_path=f"Sheet: {sheet_name}", section_type="metadata", order=order)
                )
                order += 1
        elif nrows <= SHEET_ROW_SAMPLE:
            rows = _collect_rows(ws)
            rows = trim_trailing_empty(rows)
            if not rows:
                continue
            body = rows[1:]
            rnd = random.Random(42)
            if len(body) <= 200:
                sample = rows
            else:
                first = body[:100]
                last = body[-100:]
                mid_range = list(range(100, len(body) - 100))
                pick = rnd.sample(mid_range, min(100, len(mid_range)))
                pick.sort()
                middle = [body[i] for i in pick]
                sample = [rows[0]] + first + middle + last
            md = rows_to_markdown(trim_trailing_empty(sample))
            warnings.append(
                f"Large sheet sampled: {sheet_name} ({nrows} rows; header + first 100 + 100 random middle + last 100)"
            )
            txt = md + ("\n\n" + "\n".join(comment_extra) if comment_extra else "")
            sections.append(Section(text=txt, structure_path=f"Sheet: {sheet_name}", section_type="sheet", order=order))
            order += 1
        else:
            summary_txt, sw = _summarize_large_sheet(ws, sheet_name)
            warnings.extend(sw)
            if comment_extra:
                summary_txt += "\n\n" + "\n".join(comment_extra)
            sections.append(
                Section(text=summary_txt, structure_path=f"Sheet: {sheet_name}", section_type="sheet", order=order)
            )
            order += 1

        try:
            ch = getattr(ws, "_charts", None)
            if ch:
                warnings.append(f"charts_skipped:{sheet_name}")
        except Exception:
            pass

    try:
        wb.close()
    except Exception:
        pass

    full = "\n\n".join(s.text for s in sections)
    return ExtractionResult(
        method="openpyxl",
        text=full,
        sections=sections,
        metadata=metadata,
        warnings=warnings,
    )


async def extract_xlsx(blob: bytes) -> ExtractionResult:
    return await asyncio.to_thread(extract_xlsx_sync, blob)


async def extract_xls(blob: bytes) -> ExtractionResult:
    converted = await asyncio.to_thread(libreoffice_convert, blob, ".xls", ".xlsx")
    if not converted:
        return empty_skipped("libreoffice", "conversion_failed", warnings=["xls conversion failed"])
    res = await extract_xlsx(converted)
    res.method = "libreoffice→openpyxl"
    res.warnings.insert(0, "Converted from legacy .xls via LibreOffice")
    return res
